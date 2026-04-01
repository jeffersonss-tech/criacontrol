"""
CriaControl - SIMPLES E ROBUSTO
"""
import streamlit as st
import pandas as pd
from datetime import date, datetime
import os
import uuid
import openpyxl

import auth
import database

# Configuração da página
st.set_page_config(page_title="CriaControl", page_icon="🐄", layout="wide")

# ===== SESSION STATE =====
if 'user' not in st.session_state:
    st.session_state.user = None
if 'page' not in st.session_state:
    st.session_state.page = 'login'

# ===== TENTAR LOGIN AUTOMÁTICO =====
if st.session_state.user is None:
    saved_user = database.load_session()
    if saved_user:
        st.session_state.user = saved_user
        st.session_state.page = 'dashboard'

# ===== FUNÇÕES =====
def login(username, password):
    success, user = auth.authenticate(username, password)
    if success:
        database.save_session(user)
        st.session_state.user = user
        st.session_state.page = 'dashboard'
        st.rerun()
    return success

def logout():
    database.clear_session()
    st.session_state.user = None
    st.session_state.page = 'login'
    st.rerun()

def gerar_id_automatico():
    """Gera ID automático no formato: BZ-YYYYMMDD-XXXX"""
    data = datetime.now().strftime("%Y%m%d")
    uid = str(uuid.uuid4())[:4].upper()
    return f"BZ-{data}-{uid}"

def gerar_pdf(df, titulo):
    """Gera PDF com os dados."""
    try:
        from fpdf import FPDF
    except ImportError:
        st.error("Instale: pip install fpdf")
        return
    
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    
    # Titulo
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(200, 10, txt=titulo, ln=True, align='C')
    pdf.ln(10)
    
    # Data
    pdf.set_font("Arial", size=10)
    pdf.cell(200, 10, txt=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=True, align='C')
    pdf.ln(10)
    
    # Resumo
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(200, 10, txt="Resumo", ln=True)
    pdf.set_font("Arial", size=10)
    pdf.cell(200, 8, txt=f"Total de Animais: {len(df)}", ln=True)
    pdf.cell(200, 8, txt=f"Peso Total: {df['peso_kg'].sum():.1f} kg", ln=True)
    pdf.cell(200, 8, txt=f"Media de Peso: {df['peso_kg'].mean():.1f} kg", ln=True)
    pdf.cell(200, 8, txt=f"Peso Minimo: {df['peso_kg'].min():.1f} kg", ln=True)
    pdf.cell(200, 8, txt=f"Peso Maximo: {df['peso_kg'].max():.1f} kg", ln=True)
    pdf.ln(10)
    
    # Tabela
    pdf.set_font("Arial", 'B', 10)
    pdf.cell(40, 10, "Numero", 1)
    pdf.cell(30, 10, "Lote", 1)
    pdf.cell(25, 10, "Data", 1)
    pdf.cell(20, 10, "Sexo", 1)
    pdf.cell(25, 10, "Raca", 1)
    pdf.cell(25, 10, "Peso(kg)", 1)
    pdf.ln()
    
    pdf.set_font("Arial", size=9)
    for _, row in df.head(50).iterrows():
        pdf.cell(40, 8, str(row['numero_bezerro'])[:15], 1)
        pdf.cell(30, 8, str(row['lote'])[:12], 1)
        pdf.cell(25, 8, str(row['data_pesagem']), 1)
        pdf.cell(20, 8, str(row['sexo']), 1)
        pdf.cell(25, 8, str(row['raca'])[:10], 1)
        pdf.cell(25, 8, f"{row['peso_kg']:.1f}", 1)
        pdf.ln()
    
    if len(df) > 50:
        pdf.ln(5)
        pdf.cell(200, 8, txt=f"... e mais {len(df) - 50} registros", ln=True)
    
    return pdf.output(dest='S').encode('latin-1')

def gerar_pdf_download(df, titulo):
    """Gera PDF com dados e gráficos."""
    from fpdf import FPDF
    import matplotlib.pyplot as plt
    import tempfile
    
    # Criar graficos
    fig, axes = plt.subplots(2, 2, figsize=(10, 8))
    fig.suptitle(titulo.replace('_', ' '))
    
    # Por sexo
    sexo_pesos = df.groupby('sexo')['peso_kg'].mean()
    axes[0, 0].bar(sexo_pesos.index, sexo_pesos.values)
    axes[0, 0].set_title('Media por Sexo')
    axes[0, 0].set_ylabel('Peso (kg)')
    
    # Por raca
    raca_pesos = df.groupby('raca')['peso_kg'].mean()
    axes[0, 1].bar(raca_pesos.index, raca_pesos.values)
    axes[0, 1].set_title('Media por Raca')
    axes[0, 1].set_ylabel('Peso (kg)')
    
    # Por combinacao
    combo = df.groupby(['sexo', 'raca'])['peso_kg'].mean()
    combo_labels = [f"{s} {r}" for s, r in combo.index]
    axes[1, 0].bar(combo_labels, combo.values)
    axes[1, 0].set_title('Media por Combinacao')
    axes[1, 0].set_ylabel('Peso (kg)')
    axes[1, 0].tick_params(axis='x', rotation=45)
    
    # Histograma
    axes[1, 1].hist(df['peso_kg'], bins=10, edgecolor='black')
    axes[1, 1].set_title('Distribuicao de Peso')
    axes[1, 1].set_xlabel('Peso (kg)')
    axes[1, 1].set_ylabel('Frequencia')
    
    plt.tight_layout()
    
    # Salvar em arquivo temporario
    with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
        tmp_path = tmp.name
    
    plt.savefig(tmp_path, format='png', dpi=100)
    plt.close()
    
    # Criar PDF
    pdf = FPDF()
    pdf.add_page()
    
    # Titulo
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(200, 10, txt=titulo.replace('_', ' '), ln=True, align='C')
    
    pdf.set_font("Arial", size=10)
    pdf.cell(200, 8, txt=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=True, align='C')
    pdf.ln(5)
    
    # Resumo
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(200, 10, txt="Resumo", ln=True)
    pdf.set_font("Arial", size=10)
    pdf.cell(200, 7, txt=f"Total: {len(df)} | Peso Total: {df['peso_kg'].sum():.1f} kg | Media: {df['peso_kg'].mean():.1f} kg", ln=True)
    pdf.ln(10)
    
    # Adicionar grafico
    pdf.image(tmp_path, x=10, w=190)
    pdf.ln(5)
    
    # Remover arquivo temporario
    import os
    os.remove(tmp_path)
    
    # Tabela
    pdf.set_font("Arial", 'B', 9)
    pdf.cell(35, 8, "Numero", 1)
    pdf.cell(25, 8, "Lote", 1)
    pdf.cell(22, 8, "Data", 1)
    pdf.cell(18, 8, "Sexo", 1)
    pdf.cell(22, 8, "Raca", 1)
    pdf.cell(22, 8, "Peso(kg)", 1)
    pdf.ln()
    
    pdf.set_font("Arial", size=8)
    for _, row in df.head(30).iterrows():
        pdf.cell(35, 7, str(row['numero_bezerro'])[:12], 1)
        pdf.cell(25, 7, str(row['lote'])[:10], 1)
        pdf.cell(22, 7, str(row['data_pesagem']), 1)
        pdf.cell(18, 7, str(row['sexo']), 1)
        pdf.cell(22, 7, str(row['raca'])[:9], 1)
        pdf.cell(22, 7, f"{row['peso_kg']:.1f}", 1)
        pdf.ln()
    
    if len(df) > 30:
        pdf.ln(5)
        pdf.cell(200, 7, txt=f"... e mais {len(df) - 30} registros (veja o Excel para dados completos)", ln=True)
    
    pdf_data = pdf.output(dest='S').encode('latin-1')
    filename = titulo.replace(" ", "_") + ".pdf"
    st.download_button("Baixar PDF", data=pdf_data, file_name=filename, mime="application/pdf")

# ===== PÁGINA DE LOGIN =====
def show_login():
    st.markdown("""
        <style>
        .login-box {
            max-width: 400px;
            margin: 100px auto;
            padding: 2rem;
            background: linear-gradient(135deg, #1f77b420 0%, #1f77b410 100%);
            border-radius: 1rem;
        }
        </style>
        <div class="login-box">
        <h1 style="text-align:center;">🐄 CriaControl</h1>
    """, unsafe_allow_html=True)
    
    st.markdown("**Sistema de Pesagem de Bezerros**")
    st.markdown("---")
    
    with st.form("login"):
        username = st.text_input("Usuário")
        password = st.text_input("Senha", type="password")
        submit = st.form_submit_button("🚀 Entrar")
        
        if submit:
            if login(username, password):
                st.success("Login realizado!")
            else:
                st.error("Usuário ou senha inválidos")
    
    st.markdown("---")
    st.info("💡 Admin: admin / admin123")
    
    with st.expander("Criar novo usuário"):
        new_user = st.text_input("Novo usuário")
        new_pass = st.text_input("Senha", type="password")
        if st.button("Criar"):
            success, msg = auth.create_user(new_user, new_pass)
            if success:
                st.success(f"✅ {msg}")
            else:
                st.error(f"❌ {msg}")
    
    st.markdown("</div>", unsafe_allow_html=True)
    
    # Sobre na tela de login
    st.markdown("---")
    st.markdown("""
    **ℹ️ Sobre o CriaControl v1.0**
    
    Sistema de gestão de pesagem de bezerros.
    
    **Desenvolvido por:** Jeferson Silva Santos
    
    **Contato:**
    📧 jeffersonssantos92@gmail.com
    📱 +55 82 93164-659
    🔗 GitHub: jeffersonss-tech
    
    🇧🇷 Automação para o agronegócio brasileiro
    """)

# ===== PÁGINA DASHBOARD =====
def show_dashboard():
    user = st.session_state.user
    pesagens = database.obter_pesagens(user['id'])
    stats = database.obter_estatisticas(user['id'])
    
    # Header
    col1, col2 = st.columns([3, 1])
    with col1:
        st.title("🐄 CriaControl")
        st.write(f"**Usuário:** {user['username']} ({user['role']})")
    with col2:
        if st.button("🚪 Sair"):
            logout()
    
    st.markdown("---")
    
    # Menu
    menu = st.sidebar.selectbox("Menu", ["📊 Dashboard", "📈 Relatorios", "➕ Nova Pesagem", "📋 Consultar", "👥 Gerenciar Usuários"])
    
    # ============ SOBRE ============
    st.sidebar.markdown("---")
    st.sidebar.subheader("ℹ️ Sobre")
    st.sidebar.markdown("""
    **🐄 CriaControl v1.0**
    
    Sistema de gestão de pesagem de bezerros.
    
    **Desenvolvido por:**
    Jeferson Silva Santos
    
    **Contato:**
    📧 jeffersonssantos92@gmail.com
    📱 +55 82 93164-659
    🔗 GitHub: jeffersonss-tech
    
    **Tecnologias:**
    - Streamlit + Python
    - Banco de dados isolado por usuário
    - Exportação Excel e PDF
    
    ---
    *Automação para o agronegócio brasileiro* 🇧🇷
    """)
    
    # ============ RELATORIOS ============
    if menu == "📈 Relatorios":
        st.subheader("📈 Relatorios")
        
        if not pesagens:
            st.info("Nenhuma pesagem ainda.")
        else:
            df = pd.DataFrame(pesagens)
            
            # Opcoes de relatorio
            tipo = st.radio("Tipo de Relatorio", ["Geral", "Por Lote"])
            
            if tipo == "Geral":
                st.write("### Relatorio Geral")
                
                # Estatisticas gerais
                pesos = df['peso_kg']
                st.write(f"**Total de Animais:** {len(pesos)}")
                st.write(f"**Peso Total:** {pesos.sum():.1f} kg")
                st.write(f"**Media de Peso:** {pesos.mean():.1f} kg")
                st.write(f"**Peso Minimo:** {pesos.min():.1f} kg")
                st.write(f"**Peso Maximo:** {pesos.max():.1f} kg")
                
                # Por sexo
                st.write("---")
                st.write("**Por Sexo**")
                sexo_stats = df.groupby('sexo').agg({'peso_kg': ['count', 'mean']}).round(1)
                sexo_stats.columns = ['Quantidade', 'Media']
                st.dataframe(sexo_stats)
                st.bar_chart(df.groupby('sexo')['peso_kg'].mean())
                
                # Por raca
                st.write("---")
                st.write("**Por Raca**")
                raca_stats = df.groupby('raca').agg({'peso_kg': ['count', 'mean']}).round(1)
                raca_stats.columns = ['Quantidade', 'Media']
                st.dataframe(raca_stats)
                st.bar_chart(df.groupby('raca')['peso_kg'].mean())
                
                # Por combinacao
                st.write("---")
                st.write("**Por Combinacao Sexo + Raca**")
                combo_stats = df.groupby(['sexo', 'raca']).agg({'peso_kg': ['count', 'mean']}).round(1)
                combo_stats.columns = ['Quantidade', 'Media']
                combo_stats = combo_stats.reset_index()
                combo_stats['combinacao'] = combo_stats['sexo'] + ' ' + combo_stats['raca']
                st.dataframe(combo_stats.set_index('combinacao'))
                
                combo_chart = df.groupby(['sexo', 'raca'])['peso_kg'].mean().reset_index()
                combo_chart['label'] = combo_chart['sexo'] + ' ' + combo_chart['raca']
                st.bar_chart(combo_chart.set_index('label')['peso_kg'])
                
                # Tabela completa
                st.write("---")
                with st.expander("Dados Completos"):
                    st.dataframe(df)
            
            else:
                # Por lote
                st.write("### Relatorio por Lote")
                lotes_disponiveis = ["Todos"] + sorted(df['lote'].unique().tolist())
                lote_selecionado = st.selectbox("Selecionar Lote", lotes_disponiveis)
                
                if lote_selecionado == "Todos":
                    df_lote = df
                else:
                    df_lote = df[df['lote'] == lote_selecionado]
                
                # Estatisticas do lote
                pesos_lote = df_lote['peso_kg']
                st.write(f"**Lote:** {lote_selecionado}")
                st.write(f"**Total de Animais:** {len(pesos_lote)}")
                st.write(f"**Peso Total:** {pesos_lote.sum():.1f} kg")
                st.write(f"**Media de Peso:** {pesos_lote.mean():.1f} kg")
                st.write(f"**Peso Minimo:** {pesos_lote.min():.1f} kg")
                st.write(f"**Peso Maximo:** {pesos_lote.max():.1f} kg")
                
                # Por sexo
                st.write("---")
                st.write("**Por Sexo**")
                sexo_stats = df_lote.groupby('sexo').agg({'peso_kg': ['count', 'mean']}).round(1)
                sexo_stats.columns = ['Quantidade', 'Media']
                st.dataframe(sexo_stats)
                st.bar_chart(df_lote.groupby('sexo')['peso_kg'].mean())
                
                # Por raca
                st.write("---")
                st.write("**Por Raca**")
                raca_stats = df_lote.groupby('raca').agg({'peso_kg': ['count', 'mean']}).round(1)
                raca_stats.columns = ['Quantidade', 'Media']
                st.dataframe(raca_stats)
                st.bar_chart(df_lote.groupby('raca')['peso_kg'].mean())
                
                # Por combinacao
                st.write("---")
                st.write("**Por Combinacao Sexo + Raca**")
                combo_stats = df_lote.groupby(['sexo', 'raca']).agg({'peso_kg': ['count', 'mean']}).round(1)
                combo_stats.columns = ['Quantidade', 'Media']
                combo_stats = combo_stats.reset_index()
                combo_stats['combinacao'] = combo_stats['sexo'] + ' ' + combo_stats['raca']
                st.dataframe(combo_stats.set_index('combinacao'))
                
                combo_chart = df_lote.groupby(['sexo', 'raca'])['peso_kg'].mean().reset_index()
                combo_chart['label'] = combo_chart['sexo'] + ' ' + combo_chart['raca']
                st.bar_chart(combo_chart.set_index('label')['peso_kg'])
                
                # Tabela do lote
                st.write("---")
                with st.expander("Dados do Lote"):
                    st.dataframe(df_lote)
            
            # ============ EXPORTAR ============
            st.markdown("---")
            st.write("### Exportar Dados")
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.write("**Excel**")
                if tipo == "Geral":
                    excel_df = df
                    filename = "relatorio_geral.xlsx"
                else:
                    if lote_selecionado == "Todos":
                        excel_df = df
                        filename = "relatorio_todos_lotes.xlsx"
                    else:
                        excel_df = df_lote
                        filename = f"relatorio_{lote_selecionado}.xlsx"
                
                # Exportar para Excel
                import io
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    excel_df.to_excel(writer, index=False, sheet_name='Dados')
                buffer.seek(0)
                st.download_button(
                    "Baixar Excel",
                    data=buffer,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            with col2:
                st.write("**PDF**")
                if st.button("Gerar PDF"):
                    if tipo == "Geral":
                        gerar_pdf_download(df, "Relatorio_Geral")
                    else:
                        if lote_selecionado == "Todos":
                            gerar_pdf_download(df, "Relatorio_Todos_Lotes")
                        else:
                            gerar_pdf_download(df_lote, f"Relatorio_{lote_selecionado}")
    
    # ============ DASHBOARD ============
    if menu == "📊 Dashboard":
        st.subheader("📊 Dashboard")
        
        if not pesagens:
            st.info("Nenhuma pesagem ainda. Vá em 'Nova Pesagem' para começar!")
        else:
            df = pd.DataFrame(pesagens)
            
            # ============ ESTATÍSTICAS GERAIS ============
            st.write("### Estatisticas Gerais")
            
            pesos = df['peso_kg']
            c1, c2, c3, c4, c5 = st.columns(5)
            c1.metric("Total", len(pesos))
            c2.metric("Peso Total", f"{pesos.sum():.1f} kg")
            c3.metric("Media", f"{pesos.mean():.1f} kg")
            c4.metric("Minimo", f"{pesos.min():.1f} kg")
            c5.metric("Maximo", f"{pesos.max():.1f} kg")
            
            st.markdown("---")
            
            # ============ POR SEXO ============
            st.write("### Por Sexo")
            sexo_df = df.groupby('sexo').agg({'peso_kg': ['count', 'mean']}).round(1)
            sexo_df.columns = ['Quantidade', 'Media']
            st.dataframe(sexo_df, use_container_width=True)
            
            st.bar_chart(df.groupby('sexo')['peso_kg'].mean())
            
            st.markdown("---")
            
            # ============ POR RACA ============
            st.write("### Por Raca")
            raca_df = df.groupby('raca').agg({'peso_kg': ['count', 'mean']}).round(1)
            raca_df.columns = ['Quantidade', 'Media']
            st.dataframe(raca_df, use_container_width=True)
            
            st.bar_chart(df.groupby('raca')['peso_kg'].mean())
            
            st.markdown("---")
            
            # ============ POR COMBINACAO ============
            st.write("### Por Combinacao Sexo + Raca")
            combo_df = df.groupby(['sexo', 'raca']).agg({'peso_kg': ['count', 'mean']}).round(1)
            combo_df.columns = ['Quantidade', 'Media']
            combo_df = combo_df.reset_index()
            combo_df['combinacao'] = combo_df['sexo'] + ' ' + combo_df['raca']
            st.dataframe(combo_df.set_index('combinacao'), use_container_width=True)
            
            # Grafico separado
            combo_chart = df.groupby(['sexo', 'raca'])['peso_kg'].mean().reset_index()
            combo_chart['label'] = combo_chart['sexo'] + ' ' + combo_chart['raca']
            st.bar_chart(combo_chart.set_index('label')['peso_kg'])
            
            st.markdown("---")
            
            # ============ POR LOTE ============
            st.write("### Por Lote")
            lote_df = df.groupby('lote').agg({'peso_kg': ['count', 'mean', 'min', 'max']}).round(1)
            lote_df.columns = ['Qtd', 'Media', 'Min', 'Max']
            st.dataframe(lote_df, use_container_width=True)
            
            col1, col2 = st.columns(2)
            with col1:
                st.bar_chart(df.groupby('lote').size())
            with col2:
                st.bar_chart(df.groupby('lote')['peso_kg'].mean())
            
            st.markdown("---")
            
            # ============ TABELA COMPLETA ============
            with st.expander("Ver todos os registros"):
                st.dataframe(df, use_container_width=True)
    
    # ============ NOVA PESAGEM ============
    elif menu == "➕ Nova Pesagem":
        st.subheader("➕ Nova Pesagem")

        # --- Session state para preenchimento rápido ---
        if 'np_lote' not in st.session_state:
            st.session_state.np_lote = "Novo Lote"
        if 'np_sexo' not in st.session_state:
            st.session_state.np_sexo = "Macho"
        if 'np_raca' not in st.session_state:
            st.session_state.np_raca = "Zebuinos"
        if 'np_auto_id' not in st.session_state:
            st.session_state.np_auto_id = True
        if 'np_peso' not in st.session_state:
            st.session_state.np_peso = 50.0
        if 'np_obs' not in st.session_state:
            st.session_state.np_obs = ""
        if 'np_numero' not in st.session_state:
            st.session_state.np_numero = ""
        if 'np_novo_lote' not in st.session_state:
            st.session_state.np_novo_lote = ""

        lotes = database.obter_lotes(user['id'])

        # ===== BLOCO 1: SELETOR DE LOTE (sempre primeiro) =====
        st.markdown("### 📋 Lote")
        col_lote, col_refresh = st.columns([4, 1])
        with col_lote:
            novo_ou_existente = st.radio(
                "Lote", ["Existente", "Novo Lote"],
                index=0 if st.session_state.np_lote != "Novo Lote" else 1,
                horizontal=True, label_visibility="collapsed"
            )
        with col_refresh:
            st.write("")
            if st.button("🔄", help="Atualizar lista de lotes"):
                st.rerun()

        if novo_ou_existente == "Existente":
            lotes_options = ["(selecione)"] + lotes if lotes else ["(nenhum)"]
            idx = lotes_options.index(st.session_state.np_lote) if st.session_state.np_lote in lotes_options else 0
            lote_selecionado = st.selectbox("Selecionar Lote", lotes_options, index=idx)
        else:
            lote_text_input = st.text_input(
                "Nome do novo lote",
                value=st.session_state.np_novo_lote,
                placeholder="Ex: LOTE 01",
            )
            lote_selecionado = lote_text_input.strip() if lote_text_input.strip() else "Novo Lote"

        # Determina se lote é válido para destravar o formulário
        lote_valido = (
            lote_selecionado
            and lote_selecionado != "(selecione)"
            and lote_selecionado != "(nenhum)"
            and lote_selecionado != "Novo Lote"
            and lote_selecionado != ""
        )

        st.markdown("---")

        # ===== STATS (só se lote válido e com dados) =====
        if lote_valido:
            lote_pesagens = [p for p in pesagens if p['lote'] == lote_selecionado]
            if lote_pesagens:
                pesos_lote = [p['peso_kg'] for p in lote_pesagens]
                total = len(pesos_lote)

                m1, m2, m3, m4 = st.columns(4)
                m1.metric("Total", total)
                m2.metric("Peso Total", f"{sum(pesos_lote):.1f} kg")
                m3.metric("Média", f"{sum(pesos_lote)/total:.1f} kg")
                m4.metric("Mín / Máx", f"{min(pesos_lote):.1f} / {max(pesos_lote):.1f} kg")

                masc = [p for p in lote_pesagens if p['sexo'] == 'M']
                fem = [p for p in lote_pesagens if p['sexo'] == 'F']
                media_m = sum(p['peso_kg'] for p in masc)/len(masc) if masc else 0
                media_f = sum(p['peso_kg'] for p in fem)/len(fem) if fem else 0

                s1, s2 = st.columns(2)
                s1.metric("Machos", len(masc), f"{media_m:.1f} kg méd" if masc else None)
                s2.metric("Fêmeas", len(fem), f"{media_f:.1f} kg méd" if fem else None)

                zeb = [p for p in lote_pesagens if p['raca'] == 'Zebuinos']
                cruz = [p for p in lote_pesagens if p['raca'] == 'Cruzado']
                media_z = sum(p['peso_kg'] for p in zeb)/len(zeb) if zeb else 0
                media_c = sum(p['peso_kg'] for p in cruz)/len(cruz) if cruz else 0

                r1, r2 = st.columns(2)
                r1.metric("Zebuínos", len(zeb), f"{media_z:.1f} kg méd" if zeb else None)
                r2.metric("Cruzado", len(cruz), f"{media_c:.1f} kg méd" if cruz else None)

                def comb(m, r): return [p for p in lote_pesagens if p['sexo'] == m and p['raca'] == r]
                combos = [("MZ", 'M', 'Zebuinos'), ("MC", 'M', 'Cruzado'),
                          ("FZ", 'F', 'Zebuinos'), ("FC", 'F', 'Cruzado')]
                cols = st.columns(4)
                for i, (label, sx, rc) in enumerate(combos):
                    g = comb(sx, rc)
                    med = sum(p['peso_kg'] for p in g)/len(g) if g else 0
                    cols[i].metric(label, len(g), f"{med:.1f} kg" if g else None)

                st.markdown("---")

        # ===== BLOCO 2: FORMULÁRIO (bloqueado se lote inválido) =====
        st.markdown("### 📝 Registrar Pesagem")

        if not lote_valido:
            st.warning("⚠️ Selecione ou crie um lote acima para desbloquear o registro de pesagens.")

        with st.form("pesagem", clear_on_submit=True):
            locked = not lote_valido

            row1 = st.columns([3, 1])
            with row1[0]:
                if st.session_state.np_auto_id and not locked:
                    numero = st.text_input(
                        "Número do Bezerro *",
                        value="",
                        placeholder="ID automático ✔",
                        label_visibility="collapsed"
                    )
                elif locked:
                    numero = st.text_input(
                        "Número do Bezerro *",
                        value="",
                        placeholder="Bloqueado — selecione um lote",
                        disabled=True,
                        label_visibility="collapsed"
                    )
                else:
                    numero = st.text_input(
                        "Número do Bezerro *",
                        value=st.session_state.np_numero,
                        placeholder="Ex: BZ-001"
                    )
            with row1[1]:
                st.write("")
                auto_id_toggle = st.checkbox("Auto ID", value=st.session_state.np_auto_id, disabled=locked)

            row2 = st.columns([1, 1, 1])
            with row2[0]:
                sexo = st.selectbox(
                    "Sexo", ["Macho", "Fêmea"],
                    index=["Macho", "Fêmea"].index(st.session_state.np_sexo)
                    if st.session_state.np_sexo in ["Macho", "Fêmea"] else 0,
                    disabled=locked
                )
            with row2[1]:
                raca = st.selectbox(
                    "Raça", ["Zebuinos", "Cruzado"],
                    index=["Zebuinos", "Cruzado"].index(st.session_state.np_raca)
                    if st.session_state.np_raca in ["Zebuinos", "Cruzado"] else 0,
                    disabled=locked
                )
            with row2[2]:
                peso = st.number_input(
                    "Peso (kg) *", min_value=10.0, max_value=500.0,
                    value=float(st.session_state.np_peso), step=0.5, format="%.1f",
                    disabled=locked
                )

            row3 = st.columns([1, 3])
            with row3[0]:
                data = st.date_input("Data", date.today(), disabled=locked)
            with row3[1]:
                obs = st.text_area(
                    "Observações", value=st.session_state.np_obs,
                    placeholder="Opcional...",
                    disabled=locked,
                    label_visibility="collapsed"
                )

            st.markdown("")

            submitted = st.form_submit_button(
                "💾 Salvar Pesagem" if not locked else "🔒 Selecione um lote para salvar",
                use_container_width=True,
                type="primary",
                disabled=locked
            )

            if submitted:
                if not numero and auto_id_toggle:
                    numero = gerar_id_automatico()

                if not numero or lote_selecionado in ["(selecione)", "Novo Lote", ""]:
                    st.error("Preencha número e lote!")
                else:
                    sexo_map = {"Macho": "M", "Fêmea": "F"}
                    ok = database.adicionar_pesagem(
                        user['id'], numero, peso,
                        sexo_map[sexo], raca,
                        lote_selecionado,
                        str(data),
                        datetime.now().strftime("%H:%M:%S"),
                        obs
                    )
                    if ok:
                        st.session_state.np_sexo = sexo
                        st.session_state.np_raca = raca
                        st.session_state.np_peso = peso
                        st.session_state.np_obs = obs
                        st.session_state.np_auto_id = auto_id_toggle
                        st.session_state.np_numero = ""
                        st.session_state.np_novo_lote = ""
                        if lote_selecionado != "Novo Lote":
                            st.session_state.np_lote = lote_selecionado
                        st.rerun()
                    else:
                        st.error("Erro ao salvar. Tente novamente.")

        # ===== BLOCO 3: AÇÕES (último registro) =====
        if pesagens:
            ultimo = pesagens[0]
            st.markdown("---")
            col_a1, col_a2 = st.columns([3, 1])
            with col_a1:
                st.info(f"Último registro: **{ultimo['numero_bezerro']}** — {ultimo['peso_kg']} kg — {ultimo['lote']}")
            with col_a2:
                if st.button("🗑️ Excluir Último", use_container_width=True):
                    database.deletar_pesagem(user['id'], ultimo['id'])
                    st.rerun()
    
    # ============ CONSULTAR ============
    elif menu == "📋 Consultar":
        st.subheader("Consultar Pesagens")
        
        lotes = ["Todos"] + database.obter_lotes(user['id'])
        filtro = st.selectbox("Filtrar por Lote", lotes)
        
        if pesagens:
            df_pesagens = pd.DataFrame(pesagens)
            if filtro != "Todos":
                df_pesagens = df_pesagens[df_pesagens['lote'] == filtro]
            
            st.dataframe(df_pesagens[['numero_bezerro', 'lote', 'data_pesagem', 'sexo', 'raca', 'peso_kg']], use_container_width=True)
            
            # Deletar
            st.markdown("---")
            st.write("Deletar")
            ids = [""] + list(df_pesagens['id'])
            delete_id = st.selectbox("Selecionar", ids)
            if delete_id and st.button("Deletar"):
                if database.deletar_pesagem(user['id'], delete_id):
                    st.success("Deletado!")
                    st.rerun()
            
            # Limpar tudo
            st.markdown("---")
            if st.button("Limpar TODOS os dados"):
                if database.limpar_dados(user['id']):
                    st.success("Limpo!")
                    st.rerun()
        else:
            st.info("Nenhuma pesagem encontrada.")
    
    # ============ GERENCIAR USUARIOS ============
    elif menu == "👥 Gerenciar Usuários":
        st.subheader("Gerenciar Usuarios")
        
        if user['role'] != 'admin':
            st.error("Acesso negado. Apenas administradores.")
        else:
            usuarios = auth.get_all_users()
            
            st.write("Usuarios Cadastrados")
            df_users = pd.DataFrame(usuarios)
            st.dataframe(df_users, use_container_width=True)
            
            st.markdown("---")
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.write("Editar Usuario")
                edit_id = st.selectbox("Selecionar", [u['id'] for u in usuarios if u['id'] != user['id']], 
                                     format_func=lambda x: next((u['username'] for u in usuarios if u['id'] == x), ''))
                
                if edit_id:
                    new_role = st.selectbox("Novo papel", ["user", "admin"])
                    if st.button("Salvar"):
                        auth.update_user_role(edit_id, new_role)
                        st.success("Atualizado!")
                        st.rerun()
            
            with col2:
                st.write("Excluir Usuario")
                delete_id = st.selectbox("Selecionar para excluir", [u['id'] for u in usuarios if u['id'] != user['id']], 
                                        format_func=lambda x: next((u['username'] for u in usuarios if u['id'] == x), ''), key="del")
                
                if delete_id:
                    user_to_delete = next((u for u in usuarios if u['id'] == delete_id), None)
                    if user_to_delete:
                        if st.button("Excluir"):
                            auth.delete_user(delete_id)
                            st.success("Excluido!")
                            st.rerun()
            
            # Mudar senha de outros usuarios
            st.markdown("---")
            st.write("### Mudar Senha de Usuario")
            edit_pass_id = st.selectbox("Selecionar usuario", [u['id'] for u in usuarios if u['id'] != user['id']], 
                                       format_func=lambda x: next((u['username'] for u in usuarios if u['id'] == x), ''), key="edit_pass")
            
            if edit_pass_id:
                user_to_edit = next((u for u in usuarios if u['id'] == edit_pass_id), None)
                if user_to_edit:
                    nova_senha_user = st.text_input(f"Nova senha para {user_to_edit['username']}", type="password")
                    if st.button("Alterar Senha"):
                        auth.update_user_password(edit_pass_id, nova_senha_user)
                        st.success(f"Senha de {user_to_edit['username']} alterada!")
                        st.rerun()
            
            st.markdown("---")
            st.write("### Mudar Senha")
            with st.form("mudar_senha"):
                st.write("**Mudar sua senha**")
                senha_atual = st.text_input("Senha atual", type="password")
                nova_senha = st.text_input("Nova senha", type="password")
                confirmar_senha = st.text_input("Confirmar nova senha", type="password")
                
                if st.form_submit_button("Alterar Senha"):
                    if nova_senha != confirmar_senha:
                        st.error("Nova senha e confirmação não coincidem!")
                    elif nova_senha == "":
                        st.error("Digite uma nova senha!")
                    else:
                        # Verificar senha atual
                        from auth import authenticate
                        success, user = authenticate(user['username'], senha_atual)
                        if success:
                            auth.update_user_password(user['id'], nova_senha)
                            st.success("Senha alterada com sucesso!")
                        else:
                            st.error("Senha atual incorreta!")
            
            st.markdown("---")
            st.write("Criar Novo Usuario")
            with st.form("new_user"):
                c1, c2 = st.columns(2)
                with c1:
                    new_username = st.text_input("Nome")
                    new_password = st.text_input("Senha", type="password")
                with c2:
                    new_role = st.selectbox("Papel", ["user", "admin"])
                
                if st.form_submit_button("Criar"):
                    if new_username and new_password:
                        success, msg = auth.create_user(new_username, new_password, new_role)
                        if success:
                            st.success(msg)
                            st.rerun()
                        else:
                            st.error(msg)

# ===== MAIN =====
if st.session_state.page == 'login':
    show_login()
else:
    show_dashboard()
